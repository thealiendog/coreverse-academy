#!/usr/bin/env python3
"""
build_prompts_spreadsheet.py

Reads /tmp/missing_prompts.json and generates a .xlsx spreadsheet with
one row per missing image + a ready-to-paste Midjourney prompt.

Output: /tmp/coreverse_prompts_to_generate.xlsx
"""

import json, re, sys, os
from pathlib import Path
from collections import Counter

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system('python3 -m pip install openpyxl --quiet')
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

JSON_IN  = Path('/tmp/missing_prompts.json')
XLSX_OUT = Path('/tmp/coreverse_prompts_to_generate.xlsx')

# ── Subject ID → URL asset directory ────────────────────────────────────────
ASSET_DIR = {
    'inner-world':    'inner-world',
    'cosmos':         'cosmos',
    'history':        'history',
    'future-skills':  'future-skills',
    'money':          'money',
    'creative-arts':  'creative-arts',
    'wellness':       'wellness',
    'leadership':     'leadership',
    'languages':      'languages',
    'math':           'math',
    'ela':            'ela',
    'science':        'science',
    'social_studies': 'social-studies',
}

# ── Midjourney style tail (appended to every prompt) ────────────────────────
MJ_TAIL_1_1   = 'mystical children\'s educational app illustration, glowing ethereal style, deep navy dark background #080618, soft luminous colors, no text, digital art --ar 1:1 --style raw --v 6.1'
MJ_TAIL_16_9  = 'mystical children\'s educational app illustration, glowing ethereal style, deep navy dark background #080618, soft luminous colors, no text, digital art --ar 16:9 --style raw --v 6.1'

# ── Text-cleaning helpers ────────────────────────────────────────────────────
def clean(text):
    """Strip surrounding quotes, collapse whitespace, remove trailing punctuation."""
    t = (text or '').strip().strip('"\'`')
    t = re.sub(r'\s+', ' ', t)
    t = re.sub(r'[—–-]\s*$', '', t).strip()
    return t

def first_n_words(text, n=12):
    words = clean(text).split()
    return ' '.join(words[:n])

def first_sentence(text):
    """Return text up to the first sentence-ending punctuation."""
    t = clean(text)
    m = re.match(r'^(.*?[.!?])\s', t)
    return m.group(1).rstrip('.!?').strip() if m else first_n_words(t, 14)

def topic_kw(subject, lesson_title):
    """Short subject+topic keyword phrase for 'X fact visualized'."""
    # Use first 3 meaningful words of lesson title
    stop = {'the', 'a', 'an', 'and', 'of', 'in', 'is', 'are', 'to', 'for',
            'how', 'why', 'what', 'vs', 'vs.', 'our', 'your', 'my'}
    words = [w for w in lesson_title.split() if w.lower().rstrip(':,') not in stop]
    kw = ' '.join(words[:3]).rstrip(':,')
    return kw or subject

# ── Visual phrase extractors ─────────────────────────────────────────────────

def visual_welcome(ctx, lesson_title):
    """
    Returns a concise visual phrase for a welcome splash image.
    Prefers subtitle (more descriptive); falls back to headline or lesson_title.
    """
    headline = clean(ctx.get('headline', ''))
    subtitle = clean(ctx.get('subtitle', ''))

    if subtitle and len(subtitle.split()) >= 5:
        # Use subtitle but cap at ~14 words to keep it a phrase not a sentence
        return first_n_words(subtitle, 14)
    if headline and headline != lesson_title:
        return first_n_words(headline, 12)
    return first_n_words(lesson_title, 12)


def visual_game_fact(ctx, lesson_title):
    """
    Returns visual phrase for a FACT game card — derived from match_phrase.
    match_phrase explains WHY the card is true / what it illustrates.
    """
    mp = clean(ctx.get('match_phrase', ''))
    label = clean(ctx.get('label', ''))

    if mp:
        # match_phrase often starts with "Yes!/Correct!/Right!" — skip that opener
        mp = re.sub(r'^(Yes!|Correct!|Right!|Exactly!|True!|Absolutely!|Perfect!|Nice!|Great!)\s*', '', mp)
        return first_sentence(mp)
    return first_n_words(label, 12)


def visual_game_myth(ctx):
    """
    Returns visual phrase for a MYTH game card — derived from label (the absurd claim).
    """
    label = clean(ctx.get('label', ''))
    # Trim to ~15 words — myths are usually short statements
    return first_n_words(label, 15)


def visual_magazine(ctx, lesson_title):
    """
    Returns visual phrase for a magazine image — derived from image_caption or section_headline.
    """
    caption  = clean(ctx.get('image_caption', ''))
    headline = clean(ctx.get('section_headline', ''))

    if caption and len(caption.split()) >= 4:
        return first_n_words(caption, 14)
    if headline:
        return first_n_words(headline, 12)
    return first_n_words(lesson_title, 10)


# ── Prompt builders ──────────────────────────────────────────────────────────

def build_prompt(entry):
    img_type  = entry['image_type']
    ctx       = entry.get('context', {})
    subject   = entry['subject']
    lesson_title = entry.get('lesson_title', '')
    fact_or_myth = (ctx.get('fact_or_myth') or '').lower()

    # ── Welcome (16:9) ──────────────────────────────────────────────────────
    if img_type == 'welcome':
        vp = visual_welcome(ctx, lesson_title)
        return (
            f"a glowing scene in cosmic space — softly glowing {vp} — "
            f"luminous and inviting, welcome screen for a {topic_kw(subject, lesson_title)} lesson, "
            f"{MJ_TAIL_16_9}"
        )

    # ── Game card — MYTH (1:1) ───────────────────────────────────────────────
    if img_type == 'game' and fact_or_myth == 'myth':
        vp = visual_game_myth(ctx)
        return (
            f"a softly glowing silly scene in cosmic space — comical claim that {vp}, "
            f"luminous and obviously wrong, clearly absurd and amusing, "
            f"{MJ_TAIL_1_1}"
        )

    # ── Game card — FACT or categorization (1:1) ────────────────────────────
    if img_type == 'game':
        vp = visual_game_fact(ctx, lesson_title)
        kw = topic_kw(subject, lesson_title)
        # Use bucket label if it's descriptive (not just 'fact')
        bucket = ctx.get('fact_or_myth', '')
        bucket_note = '' if bucket.lower() in ('fact', 'myth', '') else f"{bucket} "
        return (
            f"a glowing scene in cosmic space — softly glowing {vp} — "
            f"luminous and clearly true, {bucket_note}{kw} visualized, "
            f"{MJ_TAIL_1_1}"
        )

    # ── Magazine / other (1:1) ───────────────────────────────────────────────
    vp = visual_magazine(ctx, lesson_title)
    kw = topic_kw(subject, lesson_title)
    return (
        f"a glowing scene in cosmic space — softly glowing {vp} — "
        f"luminous and educational, {kw} visualized, "
        f"{MJ_TAIL_1_1}"
    )


def aspect_ratio(img_type):
    return '16:9' if img_type == 'welcome' else '1:1'


def output_path(entry):
    sid      = entry['subject_id']
    adir     = ASSET_DIR.get(sid, sid)
    filename = entry['image_filename']
    return f"public/explorer-assets/{adir}/{filename}"


def context_summary(entry):
    """One-line readable summary of what this image illustrates."""
    ctx = entry.get('context', {})
    t   = entry['image_type']
    if t == 'welcome':
        return ' — '.join(filter(None, [ctx.get('headline', ''), ctx.get('subtitle', '')]))[:120]
    if t == 'game':
        return (ctx.get('label', '') or '') [:120]
    return (ctx.get('image_caption', '') or ctx.get('section_headline', ''))[:120]


# ── Sort key: subject alpha, then lesson num, then type order, then filename ─
TYPE_ORDER = {'welcome': 0, 'magazine': 1, 'other': 2, 'game': 3}

def sort_key(e):
    return (e['subject'], e['lesson_num'], TYPE_ORDER.get(e['image_type'], 9), e['image_filename'])


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    data = json.loads(JSON_IN.read_text())
    data.sort(key=sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Prompts'

    # ── Styles ───────────────────────────────────────────────────────────────
    navy_fill    = PatternFill('solid', fgColor='080618')
    navy_font    = Font(bold=True, color='FFFFFF', size=11)
    blue_fill    = PatternFill('solid', fgColor='BFDBFE')   # welcome
    purple_fill  = PatternFill('solid', fgColor='E9D5FF')   # magazine/other
    green_fill   = PatternFill('solid', fgColor='BBF7D0')   # game
    wrap_align   = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='top')
    top_align    = Alignment(vertical='top')

    # ── Column definitions: (header, width) ─────────────────────────────────
    COLUMNS = [
        ('#',                  5),
        ('Subject',           15),
        ('Guide',             20),
        ('Lesson #',          10),
        ('Lesson Title',      35),
        ('Image Filename',    28),
        ('Type',              12),
        ('Aspect Ratio',      10),
        ('Context',           50),
        ('Midjourney Prompt', 120),
        ('Output Path',       45),
        ('Done?',             10),
    ]
    HEADERS = [c[0] for c in COLUMNS]

    # Write header row
    ws.append(HEADERS)
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = navy_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}1'

    GUIDE_FULL = {
        'Sage':  'Sage the Deer',   'Nova':  'Nova the Owl',
        'Lyra':  'Lyra the Elephant','Byte':  'Byte the Fox',
        'Ace':   'Ace the Eagle',   'Muse':  'Muse the Butterfly',
        'Terra': 'Terra the Wolf',  'Valor': 'Valor the Lion',
        'Luna':  'Luna the Parrot', 'Remi':  'Remi the Raccoon',
        'Quill': 'Quill the Porcupine','Cosmo':'Cosmo the Otter',
        'Atlas': 'Atlas the Bear',
    }

    # ── Write data rows ───────────────────────────────────────────────────────
    type_fill = {'welcome': blue_fill, 'magazine': purple_fill, 'other': purple_fill, 'game': green_fill}
    COL = {h: i for i, h in enumerate(HEADERS, start=1)}

    for row_num, entry in enumerate(data, start=1):
        img_type = entry['image_type']
        prompt   = build_prompt(entry)
        guide_full = GUIDE_FULL.get(entry.get('guide', ''), entry.get('guide', ''))

        row_vals = [
            row_num,
            entry['subject'],
            guide_full,
            f"L{entry['lesson_num']}",
            entry.get('lesson_title', ''),
            entry['image_filename'],
            img_type,
            aspect_ratio(img_type),
            context_summary(entry),
            prompt,
            output_path(entry),
            '',   # Done? — empty checkbox
        ]
        ws.append(row_vals)
        xl_row = ws.max_row

        # Type column colour
        type_cell = ws.cell(row=xl_row, column=COL['Type'])
        type_cell.fill = type_fill.get(img_type, green_fill)

        # Prompt column: wrap text
        ws.cell(row=xl_row, column=COL['Midjourney Prompt']).alignment = wrap_align

        # Context column: wrap
        ws.cell(row=xl_row, column=COL['Context']).alignment = wrap_align

        # Aspect ratio: centred
        ws.cell(row=xl_row, column=COL['Aspect Ratio']).alignment = center_align
        ws.cell(row=xl_row, column=COL['Type']).alignment = center_align
        ws.cell(row=xl_row, column=COL['#']).alignment = center_align
        ws.cell(row=xl_row, column=COL['Lesson #']).alignment = center_align

        # Everything else: top-aligned
        for c in range(1, len(HEADERS)+1):
            if ws.cell(row=xl_row, column=c).alignment == Alignment():
                ws.cell(row=xl_row, column=c).alignment = top_align

    # ── Summary ───────────────────────────────────────────────────────────────
    wb.save(XLSX_OUT)
    print(f"\nSaved: {XLSX_OUT}  ({XLSX_OUT.stat().st_size // 1024} KB)")
    print(f"Total rows: {len(data)}")
    print()
    print('By subject:')
    for subj, count in sorted(Counter(e['subject'] for e in data).items()):
        print(f"  {subj:<20} {count}")
    print()
    print('By type:')
    for typ, count in sorted(Counter(e['image_type'] for e in data).items(), key=lambda x: -x[1]):
        print(f"  {typ:<12} {count}")

if __name__ == '__main__':
    main()
