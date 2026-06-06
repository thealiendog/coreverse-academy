#!/usr/bin/env python3
"""
Image Inventory Audit — Coreverse Explorer
Generates an Excel spreadsheet: every image referenced in every Explorer lesson,
cross-checked against files on disk in public/explorer-assets/.
"""

import os, re, sys
from pathlib import Path
from collections import defaultdict

# ── Install openpyxl if needed ───────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl…", flush=True)
    os.system("python3 -m pip install openpyxl --quiet")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

ROOT         = Path('/Users/carolinamanos/innerverse-academy/app')
DATA_DIR     = ROOT / 'src' / 'data'
PUBLIC_DIR   = ROOT / 'public'
PROMPTS_DIR  = ROOT / 'docs' / 'image-prompts'
OUTPUT       = Path('/tmp/coreverse_image_inventory.xlsx')

# ── Subject configuration ────────────────────────────────────────────────────
# file_prefix   : prefix of *_explorer_l##_screens.js files
# asset_dir     : directory name under public/explorer-assets/
# prompt_prefix : prefix of docs/image-prompts/{prefix}-l##-prompts.md files
# lesson_start  : first lesson number (innerworld has no l01)
SUBJECTS = [
    {'id': 'inner-world',    'name': 'Inner World & Consciousness', 'file_prefix': 'innerworld',      'asset_dir': 'inner-world',  'prompt_prefix': 'iw',             'lesson_start': 2},
    {'id': 'cosmos',         'name': 'Cosmos & Science',            'file_prefix': 'cosmos',           'asset_dir': 'cosmos',       'prompt_prefix': 'cosmos',         'lesson_start': 1},
    {'id': 'history',        'name': 'History & World',             'file_prefix': 'historyworld',     'asset_dir': 'history',      'prompt_prefix': 'historyworld',   'lesson_start': 1},
    {'id': 'future-skills',  'name': 'Future Skills',               'file_prefix': 'futureskills',     'asset_dir': 'future-skills','prompt_prefix': 'futureskills',   'lesson_start': 1},
    {'id': 'money',          'name': 'Money & Business',            'file_prefix': 'moneybusiness',    'asset_dir': 'money',        'prompt_prefix': 'money',          'lesson_start': 1},
    {'id': 'creative-arts',  'name': 'Creative Arts',               'file_prefix': 'creativearts',     'asset_dir': 'creative-arts','prompt_prefix': 'creativearts',   'lesson_start': 1},
    {'id': 'wellness',       'name': 'Life & Wellness',             'file_prefix': 'lifewellness',     'asset_dir': 'wellness',     'prompt_prefix': 'wellness',       'lesson_start': 1},
    {'id': 'leadership',     'name': 'Social & Leadership',         'file_prefix': 'socialleadership', 'asset_dir': 'leadership',   'prompt_prefix': 'leadership',     'lesson_start': 1},
    {'id': 'languages',      'name': 'Languages (Spanish)',         'file_prefix': 'spanish',          'asset_dir': 'spanish',      'prompt_prefix': None,             'lesson_start': 1},
    {'id': 'math',           'name': 'Mathematics',                 'file_prefix': 'math',             'asset_dir': 'math',         'prompt_prefix': 'math',           'lesson_start': 1},
    {'id': 'ela',            'name': 'English Language Arts',       'file_prefix': 'ela',              'asset_dir': 'ela',          'prompt_prefix': 'ela',            'lesson_start': 1},
    {'id': 'science',        'name': 'Science',                     'file_prefix': 'science',          'asset_dir': 'science',      'prompt_prefix': 'science',        'lesson_start': 1},
    {'id': 'social_studies', 'name': 'Social Studies',              'file_prefix': 'social_studies',   'asset_dir': 'socialstudies','prompt_prefix': 'social-studies', 'lesson_start': 1},
    {'id': 'frontier',       'name': 'Frontier & Philosophy',       'file_prefix': None,               'asset_dir': 'frontier',     'prompt_prefix': None,             'lesson_start': 1},
]

# Alternate asset paths to check when primary path is missing
ALT_ASSET_DIRS = {
    'inner-world': ['game-assets/inner-world', 'explorer-assets/inner-world'],
    'cosmos':      ['game-assets/cosmos'],
    'history':     ['explorer-assets/history', 'game-assets/history'],
    'future-skills': ['game-assets/future-skills'],
    'money':       ['game-assets/money'],
    'creative-arts': ['game-assets/creative-arts'],
    'wellness':    ['game-assets/life-wellness'],
    'leadership':  ['game-assets/social-leadership'],
    'languages':   ['game-assets/spanish'],
    'math':        ['game-assets/math'],
    'ela':         ['game-assets/ela'],
    'science':     ['game-assets/science'],
    'social_studies': ['game-assets/socialstudies'],
    'frontier':    ['game-assets/frontier'],
}


# ── Regex helpers ────────────────────────────────────────────────────────────
def qs(s):
    """Return regex for a quoted string value: 'x' or "x" or `x`."""
    return r"""['"`]""" + re.escape(s) + r"""['"`]"""

def extract_str(pattern, text, group=1, flags=re.DOTALL):
    m = re.search(pattern, text, flags)
    return m.group(group).strip() if m else ''


# ── Classify image by filename convention ────────────────────────────────────
def classify_image(filename):
    f = filename.split('/')[-1]  # strip any path prefix
    if re.match(r'l\d+-welcome\.(webp|png)$', f):
        return 'welcome'
    elif re.match(r'l\d+-s\d+-.+\.(webp|png)$', f):
        return 'magazine'
    elif re.match(r'l\d+-game.+\.(webp|png)$', f, re.IGNORECASE):
        return 'game'
    else:
        return 'other'


# ── Parse a lesson screen file ────────────────────────────────────────────────
def parse_lesson_file(src_file, subject_id, asset_dir):
    """
    Returns list of image dicts:
      filename, url_path, img_type, context, lesson_num, lesson_title, guide
    """
    text = src_file.read_text(encoding='utf-8')

    # Guide name (usually on first few lines)
    guide = extract_str(r"guide:\s*['\"`]([^'\"`]+)['\"`]", text)

    # Lesson number from filename (l01 → 1)
    m = re.search(r'_l(\d{2})_screens', src_file.name)
    lesson_num = int(m.group(1)) if m else 0

    # Lesson title — find first occurrence of title: '...' inside a lessons array
    lesson_title = extract_str(r"title:\s*['\"`]([^'\"`\n]+)['\"`]", text)

    images = []

    # ── 1. Full URL path images (magazine + welcome + history game items) ──
    # These start with /explorer-assets/ or /game-assets/
    full_path_re = re.compile(
        r"""(?:visual|image)\s*:\s*[`'"](\/(explorer-assets|game-assets)\/[^`'"]+)[`'"]""",
        re.DOTALL
    )
    # Also capture surrounding context: look for imageCaption or headline nearby
    # We'll do a windowed search for context
    for m in full_path_re.finditer(text):
        url_path = m.group(1)
        filename = url_path.split('/')[-1]
        img_type = classify_image(filename)

        # Context: look for imageCaption or headline in a window around this match
        start = max(0, m.start() - 800)
        end   = min(len(text), m.end() + 800)
        window = text[start:end]
        context = (
            extract_str(r"imageCaption:\s*['\"`]([^'\"`]+)['\"`]", window) or
            extract_str(r"headline:\s*['\"`]([^'\"`]+)['\"`]", window) or
            extract_str(r"label:\s*['\"`]([^'\"`\n]+)['\"`]", window)
        )
        images.append({
            'filename': filename,
            'url_path': url_path,
            'img_type': img_type,
            'context': context[:120] if context else '',
            'lesson_num': lesson_num,
            'lesson_title': lesson_title,
            'guide': guide,
        })

    # ── 2. Bare filename images in game items ──
    # Pattern: image: 'l##-game-something.webp' (no leading slash)
    # These appear inside items: [...] arrays
    bare_img_re = re.compile(
        r"""image\s*:\s*[`'"]([^`'"\/][^`'"]*\.(webp|png))[`'"]""",
        re.DOTALL
    )
    for m in bare_img_re.finditer(text):
        filename = m.group(1)
        img_type = classify_image(filename)
        # Infer url_path from asset_dir
        url_path = f'/explorer-assets/{asset_dir}/{filename}'

        # Context: label in the same game item
        start = max(0, m.start() - 200)
        end   = min(len(text), m.end() + 400)
        window = text[start:end]
        context = extract_str(r"label:\s*['\"`]([^'\"`\n]+)['\"`]", window) or ''

        images.append({
            'filename': filename,
            'url_path': url_path,
            'img_type': img_type,
            'context': context[:120],
            'lesson_num': lesson_num,
            'lesson_title': lesson_title,
            'guide': guide,
        })

    # Deduplicate by url_path (some images appear multiple times in text)
    seen = set()
    unique = []
    for img in images:
        key = img['url_path']
        if key not in seen:
            seen.add(key)
            unique.append(img)

    return unique


# ── Parse a Midjourney prompt file ────────────────────────────────────────────
def parse_prompt_file(md_file):
    """
    Returns dict: { filename -> prompt_text }
    """
    prompts = {}
    if not md_file.exists():
        return prompts
    text = md_file.read_text(encoding='utf-8')

    # Split on ### headers followed by .webp or .png
    # Pattern: ### some-filename.webp\n```\nPROMPT\n```
    blocks = re.split(r'\n(?=###\s+\S+\.(?:webp|png))', text)
    for block in blocks:
        header_m = re.match(r'###\s+(\S+\.(?:webp|png))', block.strip())
        if not header_m:
            continue
        filename = header_m.group(1)
        # Extract first code block
        code_m = re.search(r'```[^\n]*\n([\s\S]+?)```', block)
        if code_m:
            prompts[filename] = code_m.group(1).strip()
    return prompts


# ── Check file existence ──────────────────────────────────────────────────────
def check_existence(url_path, subject_id):
    """
    Returns (status, notes, expected_path)
    status: 'EXISTS' | 'MISSING'
    """
    # Primary: the URL path directly maps to public/
    primary = PUBLIC_DIR / url_path.lstrip('/')
    primary_str = str(primary.relative_to(ROOT / 'public'))

    if primary.exists():
        return 'EXISTS', '', primary_str

    # Check alternate dirs
    filename = url_path.split('/')[-1]
    alt_dirs = ALT_ASSET_DIRS.get(subject_id, [])
    for alt_dir in alt_dirs:
        alt_path = PUBLIC_DIR / alt_dir / filename
        if alt_path.exists():
            note = f'found in /public/{alt_dir}/ (not expected path)'
            return 'EXISTS', note, primary_str

    return 'MISSING', '', primary_str


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    all_rows = []

    for subj in SUBJECTS:
        sid       = subj['id']
        sname     = subj['name']
        prefix    = subj['file_prefix']
        asset_dir = subj['asset_dir']
        pprefix   = subj['prompt_prefix']
        l_start   = subj['lesson_start']

        print(f"\n▶  {sname} ({sid})", flush=True)

        if prefix is None:
            print(f"   No lesson files — subject not yet built", flush=True)
            continue

        # Collect prompt files for this subject
        # prompt_files: { lesson_num -> { filename -> prompt } }
        prompt_cache = {}
        if pprefix:
            for ln in range(l_start, 21):
                pad = str(ln).zfill(2)
                md = PROMPTS_DIR / f'{pprefix}-l{pad}-prompts.md'
                if md.exists():
                    prompt_cache[ln] = parse_prompt_file(md)

        for ln in range(l_start, 21):
            pad = str(ln).zfill(2)
            screen_file = DATA_DIR / f'{prefix}_explorer_l{pad}_screens.js'
            if not screen_file.exists():
                continue

            images = parse_lesson_file(screen_file, sid, asset_dir)
            prompts = prompt_cache.get(ln, {})

            for img in images:
                filename  = img['filename']
                url_path  = img['url_path']
                img_type  = img['img_type']
                context   = img['context']
                lesson_num = img['lesson_num']
                lesson_title = img['lesson_title']
                guide     = img['guide']

                # Look up prompt — try exact filename, then swap extension (.webp↔.png)
                bare = filename.split('/')[-1]
                prompt_text = (
                    prompts.get(filename) or
                    prompts.get(bare) or
                    prompts.get(re.sub(r'\.(webp|png)$', lambda m: '.png' if m.group(1)=='webp' else '.webp', bare)) or
                    ''
                )

                # Check disk existence (always authoritative)
                status, notes, expected_path = check_existence(url_path, sid)

                # Status logic:
                #   EXISTS            → file is on disk (regardless of prompt)
                #   MISSING           → file not on disk, prompt IS available to generate it
                #   NO PROMPT FOUND   → file not on disk AND no Midjourney prompt exists
                if status == 'EXISTS':
                    final_status = 'EXISTS'
                elif prompt_text:
                    final_status = 'MISSING'
                elif not pprefix or not prompt_cache.get(ln):
                    # No prompt file for this subject/lesson at all
                    final_status = 'MISSING'
                else:
                    final_status = 'NO PROMPT FOUND'

                # Prompt column: show prompt text, or reason why it's absent
                if prompt_text:
                    prompt_display = prompt_text
                elif not pprefix:
                    prompt_display = '— (no prompt files for this subject)'
                elif not prompt_cache.get(ln):
                    prompt_display = '— (no prompt file for this lesson)'
                else:
                    prompt_display = '— (image not in prompt file)'

                row = {
                    'Subject':           sname,
                    'Subject ID':        sid,
                    'Guide':             guide,
                    'Lesson #':          f'L{str(ln).zfill(2)}',
                    'Lesson Title':      lesson_title,
                    'Image Filename':    filename,
                    'Image Type':        img_type,
                    'Context':           context,
                    'Midjourney Prompt': prompt_display,
                    'Expected Path':     f'public/explorer-assets/{asset_dir}/{filename}',
                    'Status':            final_status,
                    'Notes':             notes,
                }
                all_rows.append(row)

        n_exist  = sum(1 for r in all_rows if r['Subject ID'] == sid and r['Status'] == 'EXISTS')
        n_miss   = sum(1 for r in all_rows if r['Subject ID'] == sid and r['Status'] in ('MISSING', 'NO PROMPT FOUND'))
        n_noprom = sum(1 for r in all_rows if r['Subject ID'] == sid and r['Status'] == 'NO PROMPT FOUND')
        total    = sum(1 for r in all_rows if r['Subject ID'] == sid)
        print(f"   {total} images  |  {n_exist} exist  |  {n_miss} missing  |  {n_noprom} no prompt", flush=True)

    # Sort: Subject, Lesson#, Type order, Filename
    type_order = {'welcome': 0, 'magazine': 1, 'game': 2, 'other': 3}
    all_rows.sort(key=lambda r: (
        r['Subject'],
        r['Lesson #'],
        type_order.get(r['Image Type'], 9),
        r['Image Filename']
    ))

    # ── Build Excel ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── SUMMARY sheet ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = 'Summary'

    navy_fill  = PatternFill('solid', fgColor='080618')
    navy_font  = Font(bold=True, color='FFFFFF', size=11)
    bold_font  = Font(bold=True, size=11)
    header_align = Alignment(horizontal='center', vertical='center')

    # Totals
    total_all     = len(all_rows)
    total_exists  = sum(1 for r in all_rows if r['Status'] == 'EXISTS')
    total_missing = sum(1 for r in all_rows if r['Status'] in ('MISSING', 'NO PROMPT FOUND'))
    total_noprom  = sum(1 for r in all_rows if r['Status'] == 'NO PROMPT FOUND')

    ws_sum.append(['COREVERSE EXPLORER — IMAGE INVENTORY SUMMARY'])
    ws_sum['A1'].font = Font(bold=True, size=14, color='080618')
    ws_sum.append([])
    ws_sum.append(['Metric', 'Count'])
    ws_sum.append(['Total images referenced', total_all])
    ws_sum.append(['Files EXISTS on disk',    total_exists])
    ws_sum.append(['Files MISSING from disk', total_missing])
    ws_sum.append(['No Midjourney prompt',    total_noprom])
    ws_sum.append([])

    # Per-subject breakdown
    ws_sum.append(['Subject', 'Total', 'EXISTS', 'MISSING', 'NO PROMPT', '% Complete'])
    header_row = ws_sum.max_row
    for col in range(1, 7):
        cell = ws_sum.cell(row=header_row, column=col)
        cell.fill = navy_fill
        cell.font = navy_font
        cell.alignment = header_align

    # Group rows by subject
    by_subject = defaultdict(list)
    for r in all_rows:
        by_subject[r['Subject']].append(r)

    for subj in SUBJECTS:
        sname = subj['name']
        rows  = by_subject.get(sname, [])
        if not rows:
            continue
        t = len(rows)
        e = sum(1 for r in rows if r['Status'] == 'EXISTS')
        miss = sum(1 for r in rows if r['Status'] in ('MISSING', 'NO PROMPT FOUND'))
        nop  = sum(1 for r in rows if r['Status'] == 'NO PROMPT FOUND')
        pct  = f'{round(e/t*100)}%' if t else '—'
        ws_sum.append([sname, t, e, miss, nop, pct])

    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 10
    ws_sum.column_dimensions['C'].width = 10
    ws_sum.column_dimensions['D'].width = 10
    ws_sum.column_dimensions['E'].width = 12
    ws_sum.column_dimensions['F'].width = 14

    # ── INVENTORY sheet ──────────────────────────────────────────────────────
    ws = wb.create_sheet('Inventory')

    COLUMNS = [
        'Subject', 'Subject ID', 'Guide', 'Lesson #', 'Lesson Title',
        'Image Filename', 'Image Type', 'Context', 'Midjourney Prompt',
        'Expected Path', 'Status', 'Notes'
    ]
    COL_WIDTHS = [18, 14, 16, 9, 38, 32, 10, 45, 90, 45, 13, 32]

    ws.append(COLUMNS)
    for col_idx, (col, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = navy_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}1'

    # Fills
    green_fill  = PatternFill('solid', fgColor='34D399')
    red_fill    = PatternFill('solid', fgColor='FCA5A5')
    yellow_fill = PatternFill('solid', fgColor='FCD34D')
    wrap_align  = Alignment(wrap_text=True, vertical='top')
    top_align   = Alignment(vertical='top')

    status_col  = COLUMNS.index('Status') + 1
    prompt_col  = COLUMNS.index('Midjourney Prompt') + 1

    for row_data in all_rows:
        row_vals = [row_data[c] for c in COLUMNS]
        ws.append(row_vals)
        row_idx = ws.max_row

        # Status coloring
        status_cell = ws.cell(row=row_idx, column=status_col)
        s = row_data['Status']
        if s == 'EXISTS':
            status_cell.fill = green_fill
        elif s == 'MISSING':
            status_cell.fill = red_fill
        elif s == 'NO PROMPT FOUND':
            status_cell.fill = yellow_fill

        # Prompt cell wraps
        ws.cell(row=row_idx, column=prompt_col).alignment = wrap_align

        # Other cells top-aligned
        for col_idx in range(1, len(COLUMNS) + 1):
            if col_idx != prompt_col:
                ws.cell(row=row_idx, column=col_idx).alignment = top_align

    print(f"\n✅  Writing spreadsheet to {OUTPUT}…", flush=True)
    wb.save(OUTPUT)
    print(f"    Saved: {OUTPUT}", flush=True)

    # ── Summary report ────────────────────────────────────────────────────
    print(f"\n{'='*70}")
    print(f"{'SUBJECT':<30} {'TOTAL':>6} {'EXISTS':>7} {'MISSING':>8} {'NO PROMPT':>10} {'% DONE':>7}")
    print(f"{'-'*70}")
    for subj in SUBJECTS:
        sname = subj['name']
        rows  = by_subject.get(sname, [])
        if not rows:
            print(f"  {sname:<28} {'—':>6}  (no lesson files)")
            continue
        t    = len(rows)
        e    = sum(1 for r in rows if r['Status'] == 'EXISTS')
        miss = sum(1 for r in rows if r['Status'] in ('MISSING', 'NO PROMPT FOUND'))
        nop  = sum(1 for r in rows if r['Status'] == 'NO PROMPT FOUND')
        pct  = f'{round(e/t*100)}%' if t else '—'
        print(f"  {sname:<28} {t:>6}  {e:>7}  {miss:>8}  {nop:>10}  {pct:>7}")
    print(f"{'-'*70}")
    print(f"  {'TOTAL':<28} {total_all:>6}  {total_exists:>7}  {total_missing:>8}  {total_noprom:>10}  {round(total_exists/total_all*100) if total_all else 0:>6}%")
    print(f"{'='*70}")

if __name__ == '__main__':
    main()
